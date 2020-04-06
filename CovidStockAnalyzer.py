# ============================================================================
# Crunch vs Covid data using yahoofinancials
# Author - P. Dyachenko
# =============================================================================

import pandas as pd
from yahoofinancials import YahooFinancials
from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
import sys, traceback



all_tickers = ["AAL", "DAL", "BA",  "IAG.L", "CCL", "IMMU", "RCL", "KALU", "M", "CME", "GT", "SPR", "HD"]

#"CME", "V", "MA", "GMKN.ME", "AFLT.ME", "GAZP.ME" "VTBR.ME"] """RS", "BJRI", "KALU", "CME", "V"]


# extracting stock data (historical close price) for the stocks identified
close_prices = pd.DataFrame()
crunch_beg_date = "2007-06-01"
crunch_end_date = "2011-01-01"
last_5_begin = "2015-02-01"
last_5_end = "2020-01-01"
last_year_begin = "2019-01-01"
last_year_end = "2020-03-01"
covid_beg_date = "2020-01-01"
covid_end_date = datetime.today().strftime('%Y-%m-%d')
date_stamp = datetime.today().date()
result_excel = 'CovidStockData_' + date_stamp.__str__() + '.xlsx'
attempt = 0
drop = []
measures = {}
cp_tickers = all_tickers



#KPI funcitons
def get_cagr(df):
    "function to calculate the Cumulative Annual Growth Rate of a trading strategy"
    df = df.copy()
    df["daily_ret"] = df["adjclose"].pct_change()
    df["cum_return"] = (1 + df["daily_ret"]).cumprod()
    n = len(df)/252
    cagr = (df["cum_return"][len(df)-1])**(1/n) - 1
    return cagr

def get_vola(df, is_negative):
    "function to calculate annualized volatility of a trading strategy"
    df = df.copy()
    df["daily_ret"] = df["adjclose"].pct_change()
    if is_negative:
        vol = df[df["daily_ret"] < 0]["daily_ret"].std() * np.sqrt(252)
    else:
        vol = df["daily_ret"].std() * np.sqrt(252)
    return vol

def get_max_dd(df):
    "function to calculate max drawdown"
    df = df.copy()
    df["daily_ret"] = df["adjclose"].pct_change()
    df["cum_return"] = (1 + df["daily_ret"]).cumprod()
    df["cum_roll_max"] = df["cum_return"].cummax()
    df["drawdown"] = df["cum_roll_max"] - df["cum_return"]
    df["drawdown_pct"] = df["drawdown"]/df["cum_roll_max"]
    max_dd = df["drawdown_pct"].max()
    return max_dd


# Writing data to Excel
def write_to_excel(data_frame, ticker, measures, summary):
    excel_book = openpyxl.load_workbook(result_excel)
    writer = pd.ExcelWriter(result_excel, engine='openpyxl')
    writer.book = excel_book
    data_frame.to_excel(writer, sheet_name=ticker)

    # Drawing a chart
    data_frame.reset_index().plot(figsize=(15,8), kind='line', x='Date', y=['Close Price', 'Covid/Crunch_Mean', '5Y Mean'])
    plt.savefig(ticker + '.png')
    sheet = excel_book[ticker]
    sheet.cell(row=1, column=9).value = "KPI Measures"
    i = 3
    for measure, value in measures[ticker].items():
        j = 9
        sheet.cell(row=i, column=j).value = measure
        sheet.cell(row=i, column=j+1).value = value
        i += 2 if i == 7 else 1


    chart = openpyxl.drawing.image.Image(ticker + '.png')
    sheet.add_image(chart, anchor="L1")

    # Write a summary sheet
    summary_dict = {}
    if summary:
        df_summary_temp = pd.DataFrame(measures)
        summary_dict["Largest Drawdown Covid"] = [df_summary_temp.transpose().sort_values(by='drawdown_covid', ascending=False).head(1)[:]['drawdown_covid'].to_string(), 'POS']
        summary_dict["Smallest Drawdown Covid"] = [df_summary_temp.transpose().sort_values(by='drawdown_covid', ascending=True).head(1)[:]['drawdown_covid'].to_string(), 'NEG']
        summary_dict["Highest CAGR 5Y"] = [df_summary_temp.transpose().sort_values(by='cagr_5', ascending=False).head(1)[:]['cagr_5'].to_string(), 'POS']
        summary_dict["Lowest CAGR 5Y"] = [df_summary_temp.transpose().sort_values(by='cagr_5', ascending=True).head(1)[:]['cagr_5'].to_string(), 'NEG']
        summary_dict["Highest CAGR 1Y"] = [df_summary_temp.transpose().sort_values(by='cagr_1', ascending=False).head(1)[:]['cagr_1'].to_string(), 'POS']
        summary_dict["Lowest CAGR 1Y"] = [df_summary_temp.transpose().sort_values(by='cagr_1', ascending=True).head(1)[:]['cagr_1'].to_string(), 'NEG']
        summary_dict["Highest Vola 5Y"] = [df_summary_temp.transpose().sort_values(by='vola_5', ascending=False).head(1)[:]['vola_5'].to_string(), 'NEG']
        summary_dict["Lowest Vola 5Y"] = [df_summary_temp.transpose().sort_values(by='vola_5', ascending=True).head(1)[:]['vola_5'].to_string(), 'POS']
        summary_dict["Highest Neg Vola 5Y"] = [df_summary_temp.transpose().sort_values(by='neg_vola_5', ascending=False).head(1)[:]['neg_vola_5'].to_string(), 'NEG']
        summary_dict["Lowest Neg Vola 5Y"] = [df_summary_temp.transpose().sort_values(by='neg_vola_5', ascending=True).head(1)[:]['neg_vola_5'].to_string(), 'POS']

        excel_book.create_sheet(title='KPI Summary')
        sheet = excel_book['KPI Summary']
        i = 1

        for k, (kpi, value) in enumerate(summary_dict.items()):
            j = 1 if value[1] == 'POS' else 4
            print(k.__str__() + ' i: ' + i.__str__() + ' j: ' + j.__str__())
            sheet.cell(row=i, column=j).value = kpi
            j += 1
            sheet.cell(row=i, column=j).value = value[0]
            i += 1 if k % 2 == 1  else 0

        # Cleaning up default Sheet 1 in the book
        excel_book.remove(excel_book["Sheet1"])


    excel_book.save(result_excel)
    excel_book.close()

#New File Excel file prep
with pd.ExcelWriter(result_excel, mode='a') as writer:
    pd.DataFrame().to_excel(writer)

# yahoo_financials = YahooFinancials("SAVE")
# json_obj_crunch = yahoo_financials.get_historical_price_data(crunch_beg_date, crunch_end_date, "monthly")
# ohlv_crunch = json_obj_crunch["SAVE"]['prices']

while len(cp_tickers) != 0 and attempt <= 1:
    print("-----------------")
    print("attempt number ", attempt)
    print("-----------------")
    cp_tickers = [j for j in cp_tickers if j not in drop]
    for i in range(len(cp_tickers)):
        print(i)
        print(len(cp_tickers))
        try:
            yahoo_financials = YahooFinancials(cp_tickers[i])
            json_obj_crunch = yahoo_financials.get_historical_price_data(crunch_beg_date, crunch_end_date, "daily")
            json_obj_covid = yahoo_financials.get_historical_price_data(covid_beg_date, covid_end_date, "daily")
            json_obj_last5 = yahoo_financials.get_historical_price_data(last_5_begin, last_5_end, "daily")
            json_obj_lastY = yahoo_financials.get_historical_price_data(last_year_begin, last_year_end, "daily")
            ohlv_crunch = json_obj_crunch[cp_tickers[i]]['prices']
            ohlv_covid = json_obj_covid[cp_tickers[i]]['prices']
            ohlv_last5 = json_obj_last5[cp_tickers[i]]['prices']
            ohlv_last_y = json_obj_lastY [cp_tickers[i]]['prices']
            
            #Crunch data frame coocking
            temp_crunch = pd.DataFrame(ohlv_crunch)[["formatted_date", "adjclose"]].dropna()
            sorted_crunch = temp_crunch.sort_values(by=["adjclose"], ascending=True).head(100)
            crunch_worst_mean = sorted_crunch.mean()

            # Last 5 Years data frame and measures
            temp_last5 = pd.DataFrame(ohlv_last5)[["formatted_date","adjclose"]].dropna()
            measures[cp_tickers[i]]={}
            mean_5 = temp_last5.mean()[0]
            measures[cp_tickers[i]]["mean_5"] = mean_5
            measures[cp_tickers[i]]["cagr_5"] = get_cagr(temp_last5)
            measures[cp_tickers[i]]["drawdown_5"] = get_max_dd(temp_last5)
            measures[cp_tickers[i]]["vola_5"] = get_vola(temp_last5, False)
            measures[cp_tickers[i]]["neg_vola_5"] = get_vola(temp_last5, True)

            # Last 1 Year data frame and measures
            temp_lastY = pd.DataFrame(ohlv_last_y)[["formatted_date", "adjclose"]].dropna()
            measures[cp_tickers[i]]["mean_1"] = temp_lastY.mean()[0]
            measures[cp_tickers[i]]["cagr_1"] = get_cagr(temp_lastY)
            measures[cp_tickers[i]]["drawdown_1"] = get_max_dd(temp_lastY)
            measures[cp_tickers[i]]["vola_1"] = get_vola(temp_lastY, False)
            measures[cp_tickers[i]]["neg_vola_1"] = get_vola(temp_lastY, True)

            # Covid drawdown


            # Covid data frame and measures
            temp_covid = pd.DataFrame(ohlv_covid)[["formatted_date", "adjclose"]].dropna()
            final_covid = temp_covid.sort_values(by=["formatted_date"], ascending=True)
            measures[cp_tickers[i]]["drawdown_covid"] = get_max_dd(final_covid)
            final_covid = final_covid.rename(columns={"formatted_date": "Date", "adjclose": "Close Price"})
            final_covid["Daily_Ret"] = final_covid['Close Price'].pct_change()
            final_covid['5D-Mean'] = final_covid['Close Price'].rolling(5).mean()
            final_covid['Crunch_Mean'] = crunch_worst_mean[0]
            final_covid['5Y Mean'] = mean_5
            final_covid['Covid/Crunch_Mean'] = final_covid["Close Price"] / final_covid['Crunch_Mean']
            final_covid['Covid/5Y Mean'] = final_covid["Close Price"] / final_covid['5Y Mean']
            final_covid.set_index("Date", inplace=True)
            if i == len(cp_tickers) - 1:
                write_to_excel(final_covid.round(2), cp_tickers[i], measures, True)
            else:
                write_to_excel(final_covid.round(2), cp_tickers[i], measures, False)
            drop.append(cp_tickers[i])
        except Exception as e:
            print(e)
            print(cp_tickers[i]," :failed to fetch data...retrying")
            print('-' * 60)
            traceback.print_exc(file=sys.stdout)
            print('-' * 60)
            continue
    attempt += 1


# import pandas as pd

# #
# # Dictionary with list object in values
# studentData = {
#     'adjclose' : {'m':56,'p':45,'r':59, 's': 42,'t':57},
#     'adjclose_N' : {'m':76,'p':88,'r':79, 's': 41,'t':54},
#     'adjclose_X' : {'m':7,'p':8,'r':5, 's': 41,'t':54},
#     'adjclose_Y' : {'m':4,'p':33,'r':87, 's': 41,'t':54}}
#
#     # 'age' : [34, 30, 16,  42, 57],
#     # 'city' : [54, 66, 534,  42, 57]}
# # }
# df = pd.DataFrame(studentData)
# df["2Meaan"] = df["adjclose_N"].rolling(2).sum()
# df
#


# test = df.transpose().sort_values(by='m', ascending=True).head(1)[:]['m']
# print(test.to_string())
#
#
# print(df.transpose().sort_values(by='m', ascending=True).head(1)[:]['m'])


# print(df["m"])
# measures["last5"] = {}
# measures["last5"]["mean"] = df.mean()[1]
# #
# # print(measures)
#
# print(get_vola(df, False))
#
# df = pd.DataFrame(studentData)
# print(get_vola(df, True))
#
# print(df.mean())
#
# df = pd.DataFrame(studentData)
# print(df["name"][len(df)-1])
